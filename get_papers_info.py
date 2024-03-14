from bs4 import BeautifulSoup
import re
import os
import time
import openpyxl
import requests


class GoogleScholarCrawler():
    def __init__(self) -> None:
        # 0. GFW 配置
        # MAC OS: export https_proxy=http://127.0.0.1:7890 http_proxy=http://127.0.0.1:7890 all_proxy=socks5://127.0.0.1:7890
        # Windows CMD: set http_proxy=http://127.0.0.1:7890 & set https_proxy=http://127.0.0.1:7890
        # Windows PowerShell: $Env:http_proxy="http://127.0.0.1:7890";$Env:https_proxy="http://127.0.0.1:7890"

        # 1. 修改 url：
        # eg. 获取 2023 年以来的被引：点进 CITED BY 页面，点击 Since 2023，复制 url 至此
        # eg. 获取 2024 年以来的 LLM 论文：搜索 LLM，点击 Since 2024，复制 url 至此
        self.base_url = 'https://scholar.google.com/scholar?as_ylo=2024&q=Diffusion&hl=en&as_sdt=0,5'
        
        # 2. 修改 Cookie：F12 进入控制台，刷新网页，点击 Network 选项，点击 Name 列中的第一项，右侧查看 Cookie
        # Cookie 的作用只是防止 Captcha，没有 Cookie 时偶尔也能运行，更换代理后需要重新获取 Cookie
        self.headers = {
            "Cookie": '',
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0",
            "Referer": self.base_url
        }

        # 如果爬虫被中断，修改这里以断点续传，将从 start=index 开始获取
        self.index = 0
        
        # 3. 修改爬取总数：页面顶端查看（About xxx results）
        self.n = 10

        # 4. 调整列宽：Excel 中双击列标号右侧边界自动调整列宽

        self.xlsx_name = 'google_scholar.xlsx'
        self.sheet_name = 'Diffusion'

        if os.path.exists(self.xlsx_name):
            self.wb = openpyxl.load_workbook(self.xlsx_name)
            self.sheet = self.wb[self.sheet_name] if self.sheet_name in self.wb.sheetnames else self.wb.create_sheet(self.sheet_name)
        else:
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb.active
            self.sheet.title = self.sheet_name
        
        self.sheet_header = ['Title', 'Url']
        for col, content in enumerate(self.sheet_header):
            self.sheet.cell(1, col + 1, content)

    def request_url(self, url):
        try:
            res = requests.get(url, headers=self.headers)       # 不要加 verify=False，否则当页面无法获取时会直接进入下一个页面
            if res.status_code == 200:
                return res.text
        except Exception as e:
            # Debug: HTTPSConnectionPool...(Caused by SSLError... ｜ 更换节点和 Cookie
            print(e)
            return None

    def parse_html(self, html):
        soup = BeautifulSoup(html, "html.parser")               # lxml
        list = soup.find_all("div", class_="gs_r gs_or gs_scl")
        for item in list:
            paper = item.find("h3")
            # print(paper)
            paper_label = paper.find("span", class_="gs_ct1")   # [CITATION], [PDF], [HTML]
            # print(paper_label)

            paper_title = None
            paper_url = None

            # 这样写方便扩展
            if paper_label is None:
                paper_title = paper.find("a").text
                paper_url = paper.find("a").get('href')
            elif paper_label.text == "[CITATION]":              # [CITATION]
                paper_title = paper.find_all("span")[-1].text
            else:                                               # [PDF], [HTML], ...
                paper_title = paper.find("a").text
                paper_url = paper.find("a").get('href')

            yield [paper_title, paper_url]

    def write_content_to_xlsx(self, index, item) -> None:
        for col, content in enumerate(item):
            # openpyxl 从 (1, 1) 开始，row1 作为 sheet header，所以分别偏移 2 和 1
            self.sheet.cell(index + 2, col + 1, content)

    def run(self):
        match = re.search(r'start=\d+', self.base_url)
        for start in range(self.index, self.n + 1, 10):
            time.sleep(3)
            url = re.sub(r'start=\d+', f'start={start}', self.base_url) if match else f'{self.base_url}&start={start}'
            print(url)
            try:
                html = self.request_url(url)
                # Debug: 查看是否成功获取到了网页源码
                with open('debug_web_page.html', 'w') as f:
                    f.write(html)
                items = self.parse_html(html)
                for item in items:
                    # Debug: paper_title为空 ｜ 检查 temp.html 文件，更换节点和 Cookie 或回到浏览器进行 human verification
                    print(item[0])      # paper_title
                    self.write_content_to_xlsx(self.index, item)
                    self.index += 1
                print()
            except Exception as e:
                # Debug: object of type 'NoneType' has no len() ｜ 回到浏览器进行 human verification
                print(e)
                print(f'爬虫被中断，请从 index={start} 再次爬取')
                break

        self.wb.save(self.xlsx_name)


if __name__ == '__main__':
    crawer = GoogleScholarCrawler()
    crawer.run()
